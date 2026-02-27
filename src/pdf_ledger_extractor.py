# -*- coding: utf-8 -*-
"""
PDF Ledger Extractor (sanitized/public version)
-----------------------------------------------

Goal:
- Extract semi-structured ledger lines from PDF (text selectable) and export to Excel.

Output columns:
- Date | Conciliation No | Partner | Description | Debit | Credit | Balance

Key features:
- Joins wrapped lines into a single logical record
- Separates Partner vs Description with heuristics ("trigger words")
- Cleans glued tokens and duplicated segments
- Removes footer contamination ("Final Balance", "Totals")
- Exports Excel with a Summary tab (totals + ending balance)
- Optional anonymization for sharing screenshots/logs

Disclaimer:
- This is a generalized parser. Real-world PDFs vary; adjust trigger lists as needed.
"""

from __future__ import annotations

import os
import re
import sys
import hashlib
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional

# ---------------------------
# Optional auto-install (works inside Jupyter too)
# ---------------------------
def _ensure_deps():
    import importlib
    import subprocess

    pkgs = [
        ("numpy", "numpy"),
        ("pandas", "pandas"),
        ("openpyxl", "openpyxl"),
        ("pdfplumber", "pdfplumber"),
    ]
    missing = []
    for pip_name, import_name in pkgs:
        try:
            importlib.import_module(import_name)
        except Exception:
            missing.append(pip_name)

    if missing:
        print("Missing dependencies:", ", ".join(missing))
        print("Installing with pip...\n")
        # Install into current interpreter environment
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", "pip"])
        subprocess.check_call([sys.executable, "-m", "pip", "install", *missing])
        print("\nDone.\n")

_ensure_deps()

import numpy as np
import pandas as pd
import pdfplumber

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


# ---------------------------
# Configuration (generic)
# ---------------------------

DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")
MONEY_RE = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}")

FOOTER_KILL_RE = re.compile(
    r"(Saldo\s+Final\s*:|Totais\s+de\s+Debitos\s+e\s+Cr[eé]ditos\s*:)",
    re.IGNORECASE
)

SALDO_ANT_RE = re.compile(r"Saldo\s+Anterior:\s*(-?\d{1,3}(?:\.\d{3})*,\d{2})", re.IGNORECASE)
SALDO_FIN_RE = re.compile(r"Saldo\s+Final\s*:\s*(-?\d{1,3}(?:\.\d{3})*,\d{2})", re.IGNORECASE)
TOTAIS_RE = re.compile(
    r"Totais\s+de\s+Debitos\s+e\s+Cr[eé]ditos\s*:\s*"
    r"(-?\d{1,3}(?:\.\d{3})*,\d{2})\s+(-?\d{1,3}(?:\.\d{3})*,\d{2})",
    re.IGNORECASE
)

NUM_PREFIX_RE = re.compile(r"^\d{1,3}(?:\.\d{3})+(?:\s+|$)")
ADT_RE = re.compile(r"^(ADT0|ADTO|ADT)\d*$", re.IGNORECASE)

# These are generic categories/keywords frequently seen in descriptions.
# Adjust this list for your context. Keep it generic for public repo.
HIST_ONLY_PREFIX = {
    "TRANSF", "TRANSF.", "SALDO", "RESGATE", "APLICACAO", "APLICAÇÃO", "CDB", "DI",
    "PIX", "TED", "DOC", "TARIFA", "JUROS", "INSS",
}

HIST_START_WORDS = {
    "ROYALTIES", "FOLHA", "DESPESA", "DESPESAS", "OPERACIONAL",
    "PIX", "TED", "DOC", "TRANSFER", "TRANSFERÊNCIA", "TRANSFERENCIA",
    "PAGAMENTO", "PGTO", "TARIFA", "JUROS", "MULTA",
    "REEMBOLSO", "ADIANTAMENTO", "ALUGUEL",
    "NF", "NFS", "NFE", "PARCELA",
    "INSS", "REMUNERAÇÃO", "REMUNERACAO",
    "DIREITO", "IMAGEM",
    "DISTRIBUIÇÃO", "DISTRIBUICAO", "LUCRO",
    "COMISSÃO", "COMISSAO", "DEFINITIVA",
    "PREMIAÇÃO", "PREMIACAO",
    "PARCERIA", "PROJETO",
    "COMPRA",
    "CAMAROTE", "ESTACIONAMENTO",
}

# Optional: known acronyms that are valid partners (generic examples)
ACRONYM_PARTNERS = {"FCF", "CBF", "FPF", "STJD", "CND"}


@dataclass
class ExtractConfig:
    anonymize: bool = False
    anonymize_salt: str = "public-demo"
    top_n_empty_partner: int = 30
    verbose: bool = True


def clean_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def parse_br_money(s: Optional[str]) -> float:
    if s is None:
        return np.nan
    s = str(s).strip()
    if s == "":
        return np.nan
    neg = False
    if s.startswith("-"):
        neg = True
        s = s[1:].strip()
    s = s.replace(".", "").replace(",", ".")
    try:
        v = float(s)
        return -v if neg else v
    except Exception:
        return np.nan


def strip_footer_from_text(t: str) -> str:
    t = clean_spaces(t)
    t = re.split(r"(?i)\bSaldo\s+Final\s*:", t)[0].strip()
    t = re.split(r"(?i)\bTotais\s+de\s+Debitos\s+e\s+Cr[eé]ditos\s*:", t)[0].strip()
    return clean_spaces(t)


def strip_num_prefix(t: str) -> str:
    t = clean_spaces(t)
    return clean_spaces(NUM_PREFIX_RE.sub("", t, count=1))


def fix_glued_words(text: str) -> str:
    """
    Fix common glued tokens from PDF extraction:
    - LTDADIREITO -> LTDA DIREITO
    - SAINSS -> SA INSS
    """
    t = text.replace("–", "-").replace("—", "-")
    t = re.sub(r"(LTDA\.?)(?=[A-ZÁÉÍÓÚÇ])", r"\1 ", t)
    t = re.sub(r"(S\/A|S\.A\.|S\.A|SA)(?=[A-ZÁÉÍÓÚÇ])", r"\1 ", t)
    t = re.sub(r"(\))(?=[A-ZÁÉÍÓÚÇ])", r"\1 ", t)
    t = re.sub(r"-(?=[A-ZÁÉÍÓÚÇ])", r"- ", t)
    return clean_spaces(t)


def dedupe_immediate_repeat(text: str) -> str:
    """
    Remove direct duplication:
    'FOO BAR FOO BAR' -> 'FOO BAR'
    """
    text = clean_spaces(text)
    if not text:
        return text
    toks = text.split()
    n = len(toks)

    if n % 2 == 0:
        half = n // 2
        if toks[:half] == toks[half:]:
            return clean_spaces(" ".join(toks[:half]))

    for k in range(6, 1, -1):
        if n >= 2 * k and toks[:k] == toks[k:2 * k]:
            new = toks[:k] + toks[2 * k:]
            return clean_spaces(" ".join(new))

    return text


def is_noise_header(line: str) -> bool:
    low = (line or "").lower()
    if low.startswith(("empresa:", "período", "periodo", "emissão", "emissao", "usuário", "usuario")):
        return True
    # header row often contains these tokens
    if ("dt." in low and "hist" in low and "saldo" in low):
        return True
    if low.startswith("conta corrente:"):
        return True
    return False


def is_footer_line(line: str) -> bool:
    return FOOTER_KILL_RE.search(line or "") is not None


def clean_partner(p: str) -> str:
    p = clean_spaces(p)
    p = re.sub(r"\s-\s*$", "", p).strip()
    return p


def anonymize_text(text: str, salt: str) -> str:
    """
    Stable anonymization (hash) preserving distinctness.
    """
    text = clean_spaces(text)
    if not text:
        return text
    h = hashlib.sha256((salt + "::" + text).encode("utf-8")).hexdigest()[:10]
    # keep a tiny hint of original length for readability
    return f"ANON_{h}"


def infer_partner_and_description(text: str) -> Tuple[str, str]:
    """
    Split Partner vs Description using heuristics:
    - Remove footer contamination
    - Remove numeric prefixes (sometimes appear before partner)
    - Fix glued tokens & duplicates
    - Find first "description trigger" word; everything before it becomes partner
    """
    text = strip_footer_from_text(text)
    text = strip_num_prefix(text)
    text = fix_glued_words(text)
    text = dedupe_immediate_repeat(text)
    text = clean_spaces(text)

    if not text:
        return "", ""

    if text.startswith("- "):
        return "", text

    toks = text.split()
    up = [t.upper().strip("()") for t in toks]

    # description-only patterns
    if up[0] in HIST_ONLY_PREFIX:
        return "", text

    # acronym as partner
    if up[0] in ACRONYM_PARTNERS:
        return up[0], clean_spaces(" ".join(toks[1:]))

    # find first trigger
    cut_idx = None
    for i, tok in enumerate(up):
        if tok in HIST_START_WORDS or ADT_RE.match(tok):
            cut_idx = i
            break

    # hyphen used as separator
    if cut_idx is None and "-" in toks:
        j = toks.index("-")
        if j > 0:
            cut_idx = j

    if cut_idx is not None:
        if cut_idx == 0:
            return "", text

        partner = clean_partner(" ".join(toks[:cut_idx]))
        desc = clean_spaces(" ".join(toks[cut_idx:]).lstrip("-").strip())

        # guard: partner should not start with description trigger
        if partner and partner.split()[0].upper() in HIST_START_WORDS:
            return "", clean_spaces((partner + " " + desc).strip())

        return partner, desc

    # fallback: assume partner is first 2..4 tokens
    n = min(4, len(toks))
    if n >= 2:
        return clean_partner(" ".join(toks[:n])), clean_spaces(" ".join(toks[n:]))

    return "", text


def extract_lines(pdf_path: str) -> List[str]:
    lines: List[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ""
            for l in txt.split("\n"):
                l = clean_spaces(l)
                if l:
                    lines.append(l)
    return lines


def extract_summary_from_lines(lines: List[str]) -> Dict[str, float]:
    saldo_ant = np.nan
    saldo_fin = np.nan
    tot_deb = np.nan
    tot_cred = np.nan

    for l in lines:
        l = clean_spaces(l)

        m = SALDO_ANT_RE.search(l)
        if m and np.isnan(saldo_ant):
            saldo_ant = parse_br_money(m.group(1))

        m = SALDO_FIN_RE.search(l)
        if m and np.isnan(saldo_fin):
            saldo_fin = parse_br_money(m.group(1))

        m = TOTAIS_RE.search(l)
        if m and (np.isnan(tot_deb) or np.isnan(tot_cred)):
            tot_deb = parse_br_money(m.group(1))
            tot_cred = parse_br_money(m.group(2))

    return {
        "saldo_anterior_pdf": saldo_ant,
        "saldo_final_pdf": saldo_fin,
        "total_debito_pdf": tot_deb,
        "total_credito_pdf": tot_cred,
    }


def parse_main_line(line: str) -> Optional[Dict[str, str]]:
    """
    Detect a main ledger line:
    starts with DD/MM/YYYY and ends with monetary values.
    """
    parts = line.split(" ")
    if not parts or not DATE_RE.match(parts[0]):
        return None

    dt = parts[0]
    vals = MONEY_RE.findall(line)

    # partial line: date but no money captured
    if len(vals) < 1:
        return {
            "dt": dt,
            "conc": "",
            "text": clean_spaces(" ".join(parts[1:])),
            "deb": "",
            "cred": "",
            "saldo": "",
            "partial": True,
        }

    tail = vals[-3:]
    deb = cred = saldo = ""
    if len(tail) == 3:
        deb, cred, saldo = tail
    elif len(tail) == 2:
        deb, saldo = tail[0], tail[1]
    else:
        saldo = tail[0]

    rest = line[len(dt):].strip()
    for v in [saldo, cred, deb]:
        if v and rest.endswith(v):
            rest = rest[: -len(v)].strip()

    rest_parts = rest.split(" ")
    conc = ""
    if rest_parts and rest_parts[0].isdigit():
        conc = rest_parts[0]
        rest_parts = rest_parts[1:]

    return {
        "dt": dt,
        "conc": conc,
        "text": clean_spaces(" ".join(rest_parts)),
        "deb": deb,
        "cred": cred,
        "saldo": saldo,
        "partial": False,
    }


def build_records(lines: List[str]) -> List[Dict[str, str]]:
    """
    Build logical records by joining wrapped lines.
    """
    rows: List[Dict[str, str]] = []
    cur: Optional[Dict[str, str]] = None

    for line in lines:
        if is_noise_header(line) or is_footer_line(line):
            continue

        p = parse_main_line(line)

        # continuation line (no date)
        if p is None:
            if cur is not None:
                cur["text"] = clean_spaces(cur["text"] + " " + line)
            continue

        # partial line (date but no money)
        if p["partial"]:
            if cur is None:
                cur = p
            else:
                cur["text"] = clean_spaces(cur["text"] + " " + p["text"])
            continue

        # new full record: close previous
        if cur is not None:
            rows.append(cur)
        cur = p

    if cur is not None:
        rows.append(cur)

    return rows


def compute_top_empty_partner(df: pd.DataFrame, top_n: int, cfg: ExtractConfig) -> pd.DataFrame:
    """
    Get the most frequent Description values when Partner is empty.
    Useful to validate trigger words and improve heuristics.
    """
    w = df.copy()
    mask = w["Partner"].fillna("").str.strip().eq("")
    s = w.loc[mask, "Description"].fillna("").str.strip()

    s = s[s.ne("")]
    if cfg.anonymize:
        s = s.apply(lambda x: anonymize_text(x[:80], cfg.anonymize_salt))

    top = s.value_counts().head(top_n).reset_index()
    top.columns = ["Description (empty Partner)", "Count"]
    return top


def export_excel(df_lanc: pd.DataFrame, resumo: Dict[str, float], out_xlsx: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lancamentos"

    for row in dataframe_to_rows(df_lanc, index=False, header=True):
        ws.append(row)

    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {"A": 12, "B": 12, "C": 28, "D": 70, "E": 14, "F": 14, "G": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    money_fmt = '#,##0.00'
    for col_letter in ["E", "F", "G"]:
        for cell in ws[col_letter][1:]:
            cell.number_format = money_fmt

    table_ref = f"A1:G{ws.max_row}"
    tab = Table(displayName="tblLancamentos", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)

    # Summary tab
    ws2 = wb.create_sheet("Resumo")
    ws2["A1"] = "Resumo"
    ws2["A1"].font = Font(bold=True, size=14)

    rows = [
        ("Saldo Anterior (PDF)", resumo.get("saldo_anterior_pdf")),
        ("Total Débitos (calculado)", resumo.get("total_debito_calc")),
        ("Total Créditos (calculado)", resumo.get("total_credito_calc")),
        ("Saldo Final (calculado)", resumo.get("saldo_final_calc")),
        ("Total Débitos (PDF)", resumo.get("total_debito_pdf")),
        ("Total Créditos (PDF)", resumo.get("total_credito_pdf")),
        ("Saldo Final (PDF)", resumo.get("saldo_final_pdf")),
    ]

    r = 3
    for label, value in rows:
        ws2[f"A{r}"] = label
        ws2[f"A{r}"].font = Font(bold=True)
        ws2[f"B{r}"] = value
        ws2[f"B{r}"].number_format = money_fmt
        r += 1

    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 22

    wb.save(out_xlsx)


def run(pdf_path: str, out_xlsx: str, cfg: ExtractConfig) -> Tuple[pd.DataFrame, Dict[str, float], pd.DataFrame]:
    if not os.path.isfile(pdf_path):
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    lines = extract_lines(pdf_path)
    if cfg.verbose:
        print(f"Lines extracted: {len(lines)}")

    summary_pdf = extract_summary_from_lines(lines)
    records = build_records(lines)
    if not records:
        raise RuntimeError("No records found. Check if PDF is text-selectable or layout differs significantly.")

    # build dataframe
    out = []
    for r in records:
        partner, desc = infer_partner_and_description(r["text"])
        if cfg.anonymize:
            partner = anonymize_text(partner, cfg.anonymize_salt) if partner else ""
            desc = anonymize_text(desc[:120], cfg.anonymize_salt) if desc else ""

        out.append({
            "Date": r["dt"],
            "ConciliationNo": r["conc"],
            "Partner": partner,
            "Description": desc,
            "Debit_txt": r["deb"],
            "Credit_txt": r["cred"],
            "Balance_txt": r["saldo"],
        })

    df = pd.DataFrame(out)
    df["Debit"] = df["Debit_txt"].apply(parse_br_money)
    df["Credit"] = df["Credit_txt"].apply(parse_br_money)
    df["Balance"] = df["Balance_txt"].apply(parse_br_money)

    df_export = df[["Date", "ConciliationNo", "Partner", "Description", "Debit", "Credit", "Balance"]].copy()

    total_debit_calc = df_export["Debit"].fillna(0).sum()
    total_credit_calc = df_export["Credit"].fillna(0).sum()
    saldo_final_calc = df_export["Balance"].dropna().iloc[-1] if df_export["Balance"].notna().any() else np.nan

    resumo = {
        **summary_pdf,
        "total_debito_calc": float(total_debit_calc),
        "total_credito_calc": float(total_credit_calc),
        "saldo_final_calc": float(saldo_final_calc) if not np.isnan(saldo_final_calc) else np.nan,
    }

    top_empty_partner = compute_top_empty_partner(df_export.rename(columns={"Partner":"Partner","Description":"Description"}),
                                                 top_n=cfg.top_n_empty_partner,
                                                 cfg=cfg)

    export_excel(df_export, resumo, out_xlsx)

    if cfg.verbose:
        print(f"Excel generated: {out_xlsx}")

    return df_export, resumo, top_empty_partner


if __name__ == "__main__":
    # Minimal CLI (safe for sharing)
    print("\n=== PDF Ledger Extractor (public) ===\n")
    pdf_path = input("PDF path (full path): ").strip().strip('"')
    if not pdf_path:
        raise SystemExit("No PDF path provided.")

    out_xlsx = input("Output Excel path (ENTER = same folder with _extracted.xlsx): ").strip().strip('"')
    if not out_xlsx:
        base = os.path.splitext(os.path.basename(pdf_path))[0]
        out_xlsx = os.path.join(os.path.dirname(pdf_path), f"{base}_extracted.xlsx")

    anon = input("Anonymize outputs? (y/N): ").strip().lower() == "y"
    cfg = ExtractConfig(anonymize=anon, verbose=True)

    df_export, resumo, top_empty = run(pdf_path, out_xlsx, cfg)

    print("\n--- Summary ---")
    for k, v in resumo.items():
        print(f"{k}: {v}")

    print("\n--- Top descriptions where Partner is empty (for quick validation) ---")
    print(top_empty.to_string(index=False))

    input("\nPress ENTER to close...")
